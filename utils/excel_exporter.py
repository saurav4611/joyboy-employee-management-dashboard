import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from utils.logger import get_logger

logger = get_logger("utils.excel_exporter")

class ExcelExporter:
    @staticmethod
    def _get_header_style() -> dict:
        return {
            'font': Font(name='Segoe UI', bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='1e1e2e', end_color='1e1e2e', fill_type='solid'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='dee2e6'),
                right=Side(style='thin', color='dee2e6'),
                top=Side(style='thin', color='dee2e6'),
                bottom=Side(style='thin', color='dee2e6')
            )
        }

    @staticmethod
    def _get_cell_style() -> dict:
        return {
            'font': Font(name='Segoe UI', size=10, color='313244'),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='dee2e6'),
                right=Side(style='thin', color='dee2e6'),
                top=Side(style='thin', color='dee2e6'),
                bottom=Side(style='thin', color='dee2e6')
            )
        }

    @staticmethod
    def export_employees(file_path: str, data: list[dict]) -> bool:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Employees"

            headers = ["Code", "First Name", "Last Name", "Email", "Phone", "Department", "Position", "Salary", "Hire Date", "Status"]
            ws.append(headers)

            header_style = ExcelExporter._get_header_style()
            cell_style = ExcelExporter._get_cell_style()

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_style['font']
                cell.fill = header_style['fill']
                cell.alignment = header_style['alignment']
                cell.border = header_style['border']

            for emp in data:
                row_data = [
                    emp.get("employee_code", ""),
                    emp.get("first_name", ""),
                    emp.get("last_name", ""),
                    emp.get("email", ""),
                    emp.get("phone", ""),
                    emp.get("department", ""),
                    emp.get("position", ""),
                    emp.get("salary", 0.0),
                    emp.get("hire_date", ""),
                    emp.get("status_label", "")
                ]
                ws.append(row_data)
                row_idx = ws.max_row
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = cell_style['font']
                    cell.alignment = cell_style['alignment']
                    cell.border = cell_style['border']

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            logger.info("Employee Excel exported successfully to %s", file_path)
            return True
        except Exception as e:
            logger.error("Failed to export employee Excel: %s", str(e))
            return False